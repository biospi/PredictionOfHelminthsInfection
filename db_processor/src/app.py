import math
import os
import os.path
import re
import statistics
import uuid
import math
from datetime import datetime, timedelta, time

import numpy as np
import openpyxl
import tables
from cassandra.cluster import Cluster
from ipython_genutils.py3compat import xrange
from tables import *
import os.path
from collections import defaultdict
import dateutil.relativedelta
import time
import os
import glob
import xlrd
import pandas
import sys
import pymysql
import tables
import xlrd
from ipython_genutils.py3compat import xrange
from tables import *
from functools import partial
from multiprocessing import Pool
from openpyxl import load_workbook
from pycel import ExcelCompiler
import cryptography #need to be imported or pip install cryptography


sql_db = None
MAX_ACTIVITY_COUNT_BIO = 15000


class Animal(IsDescription):
    timestamp = Int32Col()
    control_station = Int64Col()
    serial_number = Int64Col()
    signal_strength = Int16Col()
    battery_voltage = Int16Col()
    first_sensor_value = Int32Col()


class Animal2(IsDescription):
    timestamp = Int32Col()
    serial_number = Int64Col()
    signal_strength_min = Int16Col()
    signal_strength_max = Int16Col()
    battery_voltage = Int16Col()
    first_sensor_value = Int32Col()


def execute_sql_query(query, records=None, log_enabled=True):
    try:
        global sql_db
        cursor = sql_db.cursor()
        if records is not None:
            if log_enabled:
                print("SQL Query: %s" % query, records[0])
            cursor.executemany(query, records)
        else:
            if log_enabled:
                print("SQL Query: %s" % query)
            cursor.execute(query)
        rows = cursor.fetchall()
        for row in rows:
            if log_enabled:
                print("SQL Answer: %s" % row)
        return rows
    except Exception as e:
        print("Exeception occured:{}".format(e))


def sql_db_flush():
    global sql_db
    sql_db.commit()


def show_all_records_in_sql_table(table_name):
    execute_sql_query("SELECT * FROM `%s`" % table_name)


def insert_m_record_to_sql_table(table_id, records):
    print("inserting %d records to table %s" % (len(records), table_id))
    query = "INSERT INTO `" + table_id + "` (timestamp, timestamp_s, serial_number, signal_strength, battery_voltage, " \
                                         "first_sensor_value) VALUES (%s, %s, %s, %s, %s, %s)"
    execute_sql_query(query, records)


def insert_m_record_to_sql_table_(table_id, records):
    print("inserting %d records to table %s" % (len(records), table_id))
    query = "INSERT INTO `" + table_id + "` (timestamp, timestamp_s, serial_number, signal_strength_max, " \
                                         "signal_strength_min, battery_voltage, first_sensor_value) VALUES (%s, %s, %s, %s, %s, %s, %s)"
    execute_sql_query(query, records)


def insert_record_to_sql_table(table_id, timestamp, timestamp_s, serial_number_s, signal_strength_s, battery_voltage_s,
                               first_sensor_value):
    values = (timestamp, timestamp_s, serial_number_s, signal_strength_s, battery_voltage_s, first_sensor_value)
    query = "INSERT INTO `" + table_id + "` (timestamp, timestamp_s, serial_number, signal_strength, battery_voltage, " \
                                         "first_sensor_value) VALUES (%d, %s, %d, %d, %d, %d)" % values
    execute_sql_query(query)


def insert_record_to_sql_table_(table_id, timestamp, timestamp_s, serial_number, signal_strength_max,
                                signal_strength_min, battery_voltage,
                                activity_level_avg):
    values = (timestamp, timestamp_s, serial_number, signal_strength_max, signal_strength_min, battery_voltage,
              activity_level_avg)
    query = "INSERT INTO `" + table_id + "` (timestamp, timestamp_s, serial_number, signal_strength_max," \
                                         " signal_strength_min, battery_voltage, first_sensor_value) VALUES (%s, %s, %s, %s, %s, %s, %s)" % values
    execute_sql_query(query)


def drop_all_tables(db_name):
    print("drop all tables in db...")
    tables = execute_sql_query("SHOW TABLES")
    for table in tables:
        name = table["Tables_in_%s" % db_name]
        execute_sql_query("DROP TABLE `%s`" % name)


def create_sql_table(name):
    print("creating sql table %s" % name)
    execute_sql_query("CREATE TABLE `%s` ("
                      "id INT AUTO_INCREMENT PRIMARY KEY,"
                      "timestamp INT,"
                      "timestamp_s VARCHAR(255),"
                      "serial_number BIGINT,"
                      "signal_strength INT,"
                      "battery_voltage INT,"
                      "first_sensor_value BIGINT"
                      ")" % name)


def create_sql_table_(name):
    print("creating sql table %s" % name)
    execute_sql_query("CREATE TABLE `%s` ("
                      "id INT AUTO_INCREMENT PRIMARY KEY,"
                      "timestamp INT,"
                      "timestamp_s VARCHAR(255),"
                      "serial_number BIGINT,"
                      "signal_strength_min INT,"
                      "signal_strength_max INT,"
                      "battery_voltage INT,"
                      "first_sensor_value BIGINT"
                      ")" % name)


def create_and_connect_to_sql_db(db_name):
    print("CREATE DATABASE %s..." % db_name)
    # Create a connection object
    db_server_name = "localhost"
    db_user = "axel"
    db_password = "Mojjo@2015"
    char_set = "utf8mb4"
    cusror_type = pymysql.cursors.DictCursor
    global sql_db
    sql_db = pymysql.connect(host=db_server_name, user=db_user, password=db_password)
    execute_sql_query('CREATE DATABASE IF NOT EXISTS %s' % db_name)
    connect_to_sql_database(db_server_name, db_user, db_password, db_name, char_set, cusror_type)


def connect_to_sql_database(db_server_name, db_user, db_password, db_name, char_set, cusror_type):
    print("connecting to db %s..." % db_name)
    global sql_db
    sql_db = pymysql.connect(host=db_server_name, user=db_user, password=db_password,
                             db=db_name, charset=char_set, cursorclass=cusror_type)


def by_size(words, size):
    return [word for word in words if len(word) == size]


def purge_file(filename):
    print("purge %s..." % filename)
    try:
        os.remove(filename)
    except FileNotFoundError:
        print("file not found.")


def get_epoch_from_datetime(date, time):
    return int(datetime.strptime((date + " " + time), '%d/%m/%y %I:%M:%S %p').timestamp())


def add_record_to_table(table, data):
    sn = data[0]["serial_number"]
    table_row = table.row
    for record in data:

        timestamp = get_epoch_from_datetime(record["date"], record["time"])
        sn = int(sn)
        signal_strength = -1
        battery_voltage = -1
        if record['signal_strength'] is not None and re.sub("[^0-9]", "", record["signal_strength"]) != '':
            signal_strength = - int(re.sub("[^0-9]", "", record["signal_strength"]))
        if record['battery_voltage'] is not None and re.sub("[^0-9]", "", record["battery_voltage"]) != '':
            battery_voltage = int(re.sub("[^0-9]", "", record["battery_voltage"]))
        else:
            try:
                # value is sometimes strored in hex
                battery_voltage = int(record["battery_voltage"], 16)
            except (ValueError, TypeError) as ex:
                print(ex)

        first_sensor_value = int(record["first_sensor_value"])
        x_min, x_max, y_min, y_max, z_min, z_max = 0, 0, 0, 0, 0, 0
        ssv = ""
        if "second_sensor_values_xyz" in record and record["second_sensor_values_xyz"] is not None:
            ssv = str(record["second_sensor_values_xyz"])
            split = ssv.split(":")
            # print(split)
            if len(split) == 6:
                x_min, x_max, y_min, y_max, z_min, z_max = int(split[0]), int(split[1]), int(split[2]), int(
                    split[3]), int(split[4]), int(split[5])
            # print(x_min, x_max, y_min, y_max, z_min, z_max)
        table_row['timestamp'] = timestamp
        table_row['serial_number'] = int(sn)
        table_row['signal_strength'] = signal_strength
        table_row['battery_voltage'] = battery_voltage
        table_row['first_sensor_value'] = first_sensor_value
        table_row.append()


def add_record_to_table_sum(table, timestamp_f, serial_number_f, signal_strenght_max, signal_strenght_min,
                            battery_voltage_min, activity_level_avg):
    # print(timestamp_f, serial_number_f, signal_strenght_max, signal_strenght_min, battery_voltage_min, activity_level_avg)
    table_row = table.row
    table_row['timestamp'] = int(timestamp_f)
    table_row['serial_number'] = int(serial_number_f)
    table_row['signal_strength_min'] = signal_strenght_min
    table_row['signal_strength_max'] = signal_strenght_max
    table_row['battery_voltage'] = battery_voltage_min
    table_row['first_sensor_value'] = activity_level_avg
    table_row.append()


def add_record_to_table_single(table, timestamp_s, serial_number_s, signal_strength_s, battery_voltage_s,
                               activity_level_s):
    table_row = table.row
    table_row['timestamp'] = timestamp_s
    table_row['serial_number'] = serial_number_s
    table_row['signal_strength'] = signal_strength_s
    table_row['battery_voltage'] = battery_voltage_s
    table_row['first_sensor_value'] = activity_level_s
    # print(timestamp_s, serial_number_s, signal_strenght_s, battery_voltage_s, activity_level_s, table.size_in_memory)
    table_row.append()


def is_same_day(time_initial, time_next):
    dt1 = datetime.fromtimestamp(time_initial)
    dt2 = datetime.fromtimestamp(time_next)
    return dt1.day == dt2.day


def is_same_month(time_initial, time_next):
    dt1 = datetime.fromtimestamp(time_initial)
    dt2 = datetime.fromtimestamp(time_next)
    return dt1.month == dt2.month


def is_same_hour(time_initial, time_next):
    dt1 = datetime.fromtimestamp(time_initial)
    dt2 = datetime.fromtimestamp(time_next)
    return dt1.hour == dt2.hour


def get_elapsed_days(time_initial, time_next):
    dt1 = datetime.fromtimestamp(time_initial)
    dt2 = datetime.fromtimestamp(time_next)
    rd = dateutil.relativedelta.relativedelta(dt2, dt1)
    return rd.days


def get_elapsed_hours(time_initial, time_next):
    dt1 = datetime.fromtimestamp(time_initial)
    dt2 = datetime.fromtimestamp(time_next)
    rd = dateutil.relativedelta.relativedelta(dt2, dt1)
    return rd.hours


def get_elapsed_minutes(dt_1, dt_2):
    return math.ceil((dt_2 - dt_1).total_seconds() / 60)


def get_elapsed_time_string(time_initial, time_next):
    dt1 = datetime.fromtimestamp(time_initial)
    dt2 = datetime.fromtimestamp(time_next)
    rd = dateutil.relativedelta.relativedelta(dt2, dt1)
    return '%02d:%02d:%02d:%02d' % (rd.days, rd.hours, rd.minutes, rd.seconds)


def get_elapsed_time(time_initial, time_next):
    dt1 = datetime.fromtimestamp(time_initial)
    dt2 = datetime.fromtimestamp(time_next)
    rd = dateutil.relativedelta.relativedelta(dt2, dt1)
    return [rd.days, rd.hours, rd.minutes, rd.seconds]


def check_if_same_day(curr_date_time, next_date_time):
    return curr_date_time.year == next_date_time.year and curr_date_time.month == next_date_time.month and \
           curr_date_time.day == next_date_time.day and curr_date_time.hour == next_date_time.hour and \
           curr_date_time.minute == next_date_time.minute


def get_first_last_timestamp(animal_records):
    animal_records = sorted(animal_records, key=lambda x: x[0]) #make sure that the records are sorted by timestamp
    first_record_date = animal_records[0][0]
    last_record_date = animal_records[-1][0]
    first_record_date_reduced_to_first_sec = datetime.strptime(datetime.fromtimestamp(first_record_date)
                                                               .strftime("%Y-%m-%d %H:%M:00"), "%Y-%m-%d %H:%M:%S")

    # reduce minutes ex 37 -> 30 or 01 -> 00
    minutes = int(str(first_record_date_reduced_to_first_sec.minute).zfill(2)[:-1]+"0")
    first_record_date_reduced_to_first_sec = first_record_date_reduced_to_first_sec.replace(minute=minutes)

    last_record_date_reduced_to_first_sec = datetime.strptime(datetime.fromtimestamp(last_record_date)
                                                              .strftime("%Y-%m-%d %H:%M:00"), "%Y-%m-%d %H:%M:%S")
    return first_record_date_reduced_to_first_sec, last_record_date_reduced_to_first_sec


def get_list_average(l):
    return int(sum(l) / float(len(l)))


def format_farm_id(farm_id):

    if farm_id == "70091100056":
        farm_id = "cedara_" + farm_id

    if farm_id == "70091100060":
        farm_id = "bothaville_" + farm_id

    if farm_id == "70101100005":
        farm_id = "elandsberg_" + farm_id

    if farm_id == "70101100025":
        farm_id = "eenzaamheid_" + farm_id

    if farm_id == "70101100029":
        farm_id = "msinga_" + farm_id

    if farm_id == "70101200027":
        farm_id = "delmas_" + farm_id

    return farm_id


def init_database(farm_id):
    print(sys.argv)

    print("store data in sql database...")
    create_sql_table("%s_resolution_min" % farm_id)
    create_sql_table_("%s_resolution_5min" % farm_id)
    create_sql_table_("%s_resolution_10min" % farm_id)
    create_sql_table_("%s_resolution_month" % farm_id)
    create_sql_table_("%s_resolution_week" % farm_id)
    create_sql_table_("%s_resolution_day" % farm_id)
    create_sql_table_("%s_resolution_hour" % farm_id)
    return None, None, None, None, None, None, None

    if sys.argv[1] == 'h5':
        print("store data in h5 database...")
        # init new .h5 file for receiving sorted data
        FILTERS = tables.Filters(complib='blosc', complevel=9)
        compression = False
        if compression:
            # purge_file(farm_id + "_data_compressed_blosc.h5")
            h5file = tables.open_file(farm_id + "_data_compressed_blosc.h5", "w", driver="H5FD_CORE", filters=FILTERS)
        else:
            # purge_file(farm_id + ".h5")
            h5file = tables.open_file(farm_id + ".h5", "w", driver="H5FD_CORE")

        group_min = h5file.create_group("/", "resolution_min", 'resolution per min')
        group_5min = h5file.create_group("/", "resolution_5min", 'resolution per 5min')
        group_10min = h5file.create_group("/", "resolution_10min", 'resolution per 10min')
        group_m = h5file.create_group("/", "resolution_month", 'resolution per month')
        group_w = h5file.create_group("/", "resolution_week", 'resolution per week')
        group_d = h5file.create_group("/", "resolution_day", 'resolution per day')
        group_h = h5file.create_group("/", "resolution_hour", 'resolution per hour')

        table_min = h5file.create_table(group_min, "data", Animal, "Animal data activity level averaged by min")
        table_5min = h5file.create_table(group_5min, "data", Animal, "Animal data activity level averaged by 5min")
        table_10min = h5file.create_table(group_10min, "data", Animal, "Animal data activity level averaged by 10min")
        table_h = h5file.create_table(group_h, "data", Animal2, "Animal data activity level averaged by hour")
        table_d = h5file.create_table(group_d, "data", Animal2, "Animal data activity level averaged by day")
        table_w = h5file.create_table(group_w, "data", Animal2, "Animal data activity level averaged by week")
        table_m = h5file.create_table(group_m, "data", Animal2, "Animal data activity level averaged by month")

        return table_min, table_5min, table_10min, table_h, table_d, table_w, table_m


def resample_to_min(first_timestamp, last_timestamp, animal_records):
    data = []
    n_minutes_in_between = get_elapsed_minutes(first_timestamp, last_timestamp)
    structure_m = {}
    for i in xrange(0, n_minutes_in_between):
        next_timestamp = first_timestamp + timedelta(minutes=i)
        next_timestamp_human_readable = next_timestamp.strftime("%Y-%m-%dT%H:%M")
        serial_number = animal_records[0][2]  # just take first record and get serial number
        structure_m[next_timestamp_human_readable] = [int(time.mktime(next_timestamp.timetuple())),
                                                      next_timestamp_human_readable, serial_number, None, None, None]

    # fill in resampled structure with records data
    for record in animal_records:
        record_timestamp_f = datetime.fromtimestamp(record[0]).strftime("%Y-%m-%dT%H:%M")
        if record_timestamp_f in structure_m:
            struct = structure_m[record_timestamp_f]
            if struct[3] is None:
                struct[3] = []
            struct[3].append(record[3]) #signal strenght
            if struct[4] is None:
                struct[4] = []
            struct[4].append(record[4]) #battery level
            if struct[5] is None:
                struct[5] = 0
            if record[5] is not None:
                struct[5] += record[5]  # accumulate activity level if in same hour

    # filled in record data in struct save to array for save in db
    for record in structure_m.values():
        activity = record[5]
        if activity is None:
            data.append(record)
        else:
            data.append((record[0], record[1], record[2], get_list_average(record[3]), get_list_average(record[4]), activity))

    return data


def custom_round(x, base=5):
    return base * math.floor(x/base)


def resample_to_5min(first_timestamp, last_timestamp, animal_records):
    data = []
    n_minutes_in_between = int(get_elapsed_minutes(first_timestamp, last_timestamp)/5)
    structure_m = {}
    for i in xrange(0, n_minutes_in_between):
        next_timestamp = first_timestamp + timedelta(minutes=i*5)
        next_timestamp_human_readable = next_timestamp.strftime("%Y-%m-%dT%H:%M")
        serial_number = animal_records[0][2]  # just take first record and get serial number
        structure_m[next_timestamp_human_readable] = [int(time.mktime(next_timestamp.timetuple())),
                                                      next_timestamp_human_readable, serial_number, None, None, None, None]

    # fill in resampled structure with records data
    for record in animal_records:
        record_timestamp_f = datetime.fromtimestamp(record[0]).strftime("%Y-%m-%dT%H:%M")
        minutes_s = str(custom_round(int(record_timestamp_f[-2:]), base=5)).zfill(2)
        record_timestamp_f = record_timestamp_f[:-2] + minutes_s
        if record_timestamp_f in structure_m:
            struct = structure_m[record_timestamp_f]
            if struct[3] is None:
                struct[3] = []
            struct[3].append(record[3]) #signal strenght
            if struct[4] is None:
                struct[4] = []
            struct[4].append(record[4]) #battery level
            if struct[5] is None:
                struct[5] = 0
            if record[5] is not None:
                struct[5] += record[5]  # accumulate activity level if in same 10min

    # filled in record data in struct save to array for save in db
    for record in structure_m.values():
        activity = record[5]
        if activity is None:
            data.append(record)
        else:
            data.append((record[0], record[1], record[2], max(record[3]), min(record[3]), min(record[4]), activity))
    return data


def roundup(x):
    return int(math.ceil(x / 10.0)) * 10


def resample_to_10min(first_timestamp, last_timestamp, animal_records):
    data = []
    n_minutes_in_between = int(get_elapsed_minutes(first_timestamp, last_timestamp)/10)
    structure_m = {}
    for i in xrange(0, n_minutes_in_between):
        next_timestamp = first_timestamp + timedelta(minutes=i*10)
        next_timestamp_human_readable = next_timestamp.strftime("%Y-%m-%dT%H:%M")
        serial_number = animal_records[0][2]  # just take first record and get serial number
        structure_m[next_timestamp_human_readable] = [int(time.mktime(next_timestamp.timetuple())),
                                                      next_timestamp_human_readable, serial_number, None, None, None, None]
    # fill in resampled structure with records data
    for record in animal_records:
        record_timestamp_f = datetime.fromtimestamp(record[0]).strftime("%Y-%m-%dT%H:%M")
        minutes_s = str(custom_round(int(record_timestamp_f[-2:]), base=10)).zfill(2)
        record_timestamp_f = record_timestamp_f[:-2] + minutes_s

        if record_timestamp_f in structure_m:
            struct = structure_m[record_timestamp_f]
            if struct[3] is None:
                struct[3] = []
            struct[3].append(record[3]) #signal strenght
            if struct[4] is None:
                struct[4] = []
            struct[4].append(record[4]) #battery level
            if struct[5] is None:
                struct[5] = 0
            if record[5] is not None:
                struct[5] += record[5]  # accumulate activity level if in same 10min

    # filled in record data in struct save to array for save in db
    for record in structure_m.values():
        activity = record[5]
        if activity is None:
            data.append(record)
        else:
            data.append((record[0], record[1], record[2], max(record[3]), min(record[3]), min(record[4]), activity))
    return data


def resample_to_hour(first_timestamp, last_timestamp, animal_records):
    data = []
    n_hours_in_between = int(get_elapsed_minutes(first_timestamp, last_timestamp) / 60)
    structure_m = {}
    for i in xrange(0, n_hours_in_between):
        next_timestamp = first_timestamp + timedelta(hours=i)
        next_timestamp_human_readable = next_timestamp.strftime("%Y-%m-%dT%H:00")
        serial_number = animal_records[0][2] #just take first record and get serial number
        structure_m[next_timestamp_human_readable] = [int(time.mktime(next_timestamp.timetuple())),
                                                      next_timestamp_human_readable, serial_number, None, None, None, None]

    # fill in resampled structure with records data
    for record in animal_records:
        record_timestamp_f = datetime.fromtimestamp(record[0]).strftime("%Y-%m-%dT%H:00")
        if record_timestamp_f in structure_m:
            struct = structure_m[record_timestamp_f]
            if struct[3] is None:
                struct[3] = []
            struct[3].append(record[3]) #signal strenght
            if struct[4] is None:
                struct[4] = []
            struct[4].append(record[4]) #battery level
            if struct[5] is None:
                struct[5] = 0
            if record[5] is not None:
                struct[5] += record[5]  # accumulate activity level if in same hour

    # filled in record data in struct save to array for save in db
    for record in structure_m.values():
        activity = record[5]
        if activity is None:
            data.append(record)
        else:
            data.append((record[0], record[1], record[2], max(record[3]), min(record[3]), min(record[4]), activity))
    return data


def resample_to_day(first_timestamp, last_timestamp, animal_records):
    data = []
    n_days_in_between = int(get_elapsed_minutes(first_timestamp, last_timestamp)/60/24)
    structure_m = {}
    #build the time structure
    for i in xrange(0, n_days_in_between):
        next_timestamp = first_timestamp + timedelta(days=i)
        next_timestamp_human_readable = next_timestamp.strftime("%Y-%m-%dT00:00")
        serial_number = animal_records[0][2] #just take first record and get serial number
        structure_m[next_timestamp_human_readable] = [int(time.mktime(next_timestamp.timetuple())),
                                                      next_timestamp_human_readable, serial_number, None, None, None, None]

    # fill in resampled structure with records data
    for record in animal_records:
        record_timestamp_f = datetime.fromtimestamp(record[0]).strftime("%Y-%m-%dT00:00")
        if record_timestamp_f in structure_m:
            struct = structure_m[record_timestamp_f]
            if struct[3] is None:
                struct[3] = []
            struct[3].append(record[3]) #signal strenght
            if struct[4] is None:
                struct[4] = []
            struct[4].append(record[4]) #battery level
            if struct[5] is None:
                struct[5] = 0
            if record[5] is not None:
                struct[5] += record[5]  # accumulate activity level if in same hour

    # filled in record data in struct save to array for save in db
    for record in structure_m.values():
        activity = record[5]
        if activity is None:
            data.append(record)
        else:
            data.append((record[0], record[1], record[2], max(record[3]), min(record[3]), min(record[4]), activity))
    return data


def resample_to_week(first_timestamp, last_timestamp, animal_records):
    data = []
    n_weeks_in_between = int(math.ceil(get_elapsed_minutes(first_timestamp, last_timestamp)/60/24/7))
    structure_m = {}
    #build the time structure
    for i in xrange(0, n_weeks_in_between):
        next_timestamp = first_timestamp + timedelta(days=i*7)
        next_timestamp_human_readable = next_timestamp.strftime("%Y-%m-%dT00:00")
        serial_number = animal_records[0][2] #just take first record and get serial number
        structure_m[next_timestamp_human_readable] = [int(time.mktime(next_timestamp.timetuple())),
                                                      next_timestamp_human_readable, serial_number, None, None, None, None]

    # fill in resampled structure with records data
    for record in animal_records:
        record_timestamp_f = datetime.fromtimestamp(record[0]).strftime("%Y-%m-%dT00:00")
        if record_timestamp_f in structure_m:
            struct = structure_m[record_timestamp_f]
            if struct[3] is None:
                struct[3] = []
            struct[3].append(record[3]) #signal strenght
            if struct[4] is None:
                struct[4] = []
            struct[4].append(record[4]) #battery level
            if struct[5] is None:
                struct[5] = 0
            if record[5] is not None:
                struct[5] += record[5]  # accumulate activity level if in same hour

    # filled in record data in struct save to array for save in db
    for record in structure_m.values():
        activity = record[5]
        if activity is None:
            data.append(record)
        else:
            data.append((record[0], record[1], record[2], max(record[3]), min(record[3]), min(record[4]), activity))
    return data


def resample_to_month(first_timestamp, last_timestamp, animal_records):
    data = []
    n_months_in_between = int(math.ceil(get_elapsed_minutes(first_timestamp, last_timestamp)/60/24/7/30))
    structure_m = {}
    #build the time structure
    for i in xrange(0, n_months_in_between):
        next_timestamp = first_timestamp + timedelta(days=i*30)
        next_timestamp_human_readable = next_timestamp.strftime("%Y-%m-%dT00:00")
        serial_number = animal_records[0][2] #just take first record and get serial number
        structure_m[next_timestamp_human_readable] = [int(time.mktime(next_timestamp.timetuple())),
                                                      next_timestamp_human_readable, serial_number, None, None, None, None]

    # fill in resampled structure with records data
    for record in animal_records:
        record_timestamp_f = datetime.fromtimestamp(record[0]).strftime("%Y-%m-%dT00:00")
        if record_timestamp_f in structure_m:
            struct = structure_m[record_timestamp_f]
            if struct[3] is None:
                struct[3] = []
            struct[3].append(record[3]) #signal strenght
            if struct[4] is None:
                struct[4] = []
            struct[4].append(record[4]) #battery level
            if struct[5] is None:
                struct[5] = 0
            if record[5] is not None:
                struct[5] += record[5]  # accumulate activity level if in same hour

    # filled in record data in struct save to array for save in db
    for record in structure_m.values():
        activity = record[5]
        if activity is None:
            data.append(record)
        else:
            data.append((record[0], record[1], record[2], max(record[3]), min(record[3]), min(record[4]), activity))
    return data


def process_raw_h5files(path):
    print(path)
    h5_raw = tables.open_file(path, "r")
    data = h5_raw.root.table
    list_raw = []
    print("loading data...")
    for idx, x in enumerate(data):  # need idx for data iteration?
        farm_id = x['control_station']
        # if farm_id != 70101200027: #todo remove
        #     continue
        value = (x['timestamp'], farm_id, x['serial_number'], x['signal_strength'], x['battery_voltage'],
                 x['first_sensor_value'] if 0 < x['first_sensor_value'] < MAX_ACTIVITY_COUNT_BIO else None,
                 datetime.fromtimestamp(x['timestamp']).strftime("%Y-%m-%dT%H:%M:%S"))
        list_raw.append(value)
    # group records by farm id/control_station
    groups = defaultdict(list)
    for i, obj in enumerate(list_raw):
        groups[obj[1]].append(obj)
    animal_list_grouped_by_farmid = list(groups.values())

    for group in animal_list_grouped_by_farmid:
        farm_id = str(group[0][1])
        process_raw_file(farm_id, group)


def create_mean_median_animal_(data):
    print('create_mean_median_animal...')
    # group records by epoch
    groups = defaultdict(list)
    for i, obj in enumerate(data):
        groups[obj[1]].append(obj)
    grouped_by_epoch = list(groups.values())
    mean_data = []
    median_data = []
    for item in grouped_by_epoch:
        mean_activity = {}
        mean_battery = 0
        mean_signal_strengh = 0
        epoch = item[0][0]
        epoch_f = item[0][1]
        mean_serial_number = 50000000000
        median_serial_number = 60000000000
        median_activity = {}
        median_battery = {}
        median_signal_strengh = {}
        for record in item:
            median_activity[record[2]] = record[5] if record[5] is not None else 0
            median_battery[record[2]] = record[4] if record[4] is not None else 0
            median_signal_strengh[record[2]] = record[3] if record[3] is not None else 0

            if record[1] not in mean_activity:
                mean_activity[record[1]] = []

            if record[5] is not None and record[5] > 0:
                mean_activity[record[1]].append(record[5])

            mean_battery += record[4] if record[4] is not None else 0
            mean_signal_strengh += record[3] if record[3] is not None else 0

        ss = int(mean_signal_strengh/len(item))
        bat = int(mean_battery/len(item))
        m_a_l = list(mean_activity.values())[0]
        m_a = None if len(m_a_l) == 0 else sum(m_a_l)/len(m_a_l)
        mean_record = (epoch, epoch_f, mean_serial_number, ss if ss != 0 else None,
                       bat if bat != 0 else None, m_a)
        mean_data.append(mean_record)

        median_a = statistics.median(sorted(list(median_activity.values())))
        median_b = statistics.median(sorted(list(median_battery.values())))
        median_ss = statistics.median(sorted(list(median_signal_strengh.values())))
        median_record = (epoch, epoch_f, median_serial_number, median_ss if median_ss != 0 else None, median_b if median_b != 0 else None, median_a if median_b != 0 else None)
        median_data.append(median_record)
    result = mean_data + median_data
    return result


def create_indexes(farm_id):
    execute_sql_query("CREATE INDEX ix__%s_resolution_month__sn__ts on %s_resolution_month(serial_number, timestamp, first_sensor_value )" % (farm_id, farm_id), log_enabled=True)
    execute_sql_query("CREATE INDEX ix__%s_resolution_week__sn__ts on %s_resolution_week(serial_number, timestamp, first_sensor_value )" % (farm_id, farm_id), log_enabled=True)
    execute_sql_query("CREATE INDEX ix__%s_resolution_day__sn__ts on %s_resolution_day(serial_number, timestamp, first_sensor_value )" % (farm_id, farm_id), log_enabled=True)
    execute_sql_query("CREATE INDEX ix__%s_resolution_hour__sn__ts on %s_resolution_hour(serial_number, timestamp, first_sensor_value )" % (farm_id, farm_id), log_enabled=True)
    execute_sql_query("CREATE INDEX ix__%s_resolution_10min__sn__ts on %s_resolution_10min(serial_number, timestamp, first_sensor_value )" % (farm_id, farm_id), log_enabled=True)
    execute_sql_query("CREATE INDEX ix__%s_resolution_5min__sn__ts on %s_resolution_5min(serial_number, timestamp, first_sensor_value )" % (farm_id, farm_id), log_enabled=True)
    execute_sql_query("CREATE INDEX ix__%s_resolution_min__sn__ts on %s_resolution_min(serial_number, timestamp, first_sensor_value )" % (farm_id, farm_id), log_enabled=True)
    

def create_mean_median_animal(data):
    print('create_mean_median_animal...')
    # group records by epoch
    groups = defaultdict(list)
    for i, obj in enumerate(data):
        groups[obj[1]].append(obj)
    grouped_by_epoch = list(groups.values())
    mean_data = []
    median_data = []

    for item in grouped_by_epoch:
        mean_activity = {}
        mean_battery = 0
        mean_max_signal_strengh = 0
        mean_min_signal_strengh = 0
        epoch = item[0][0]
        epoch_f = item[0][1]
        mean_serial_number = 50000000000
        median_serial_number = 60000000000
        median_activity = {}
        median_battery = {}
        median_max_signal_strengh = {}
        median_min_signal_strengh = {}
        for record in item:
            median_activity[record[2]] = record[6] if record[6] is not None else 0
            median_battery[record[2]] = record[5] if record[5] is not None else 0
            median_max_signal_strengh[record[2]] = record[4] if record[4] is not None else 0
            median_min_signal_strengh[record[2]] = record[3] if record[3] is not None else 0

            if record[1] not in mean_activity:
                mean_activity[record[1]] = []

            if record[6] is not None and record[6] > 0:
                mean_activity[record[1]].append(record[6])

            mean_battery += record[5] if record[5] is not None else 0
            mean_max_signal_strengh += record[4] if record[4] is not None else 0
            mean_min_signal_strengh += record[3] if record[3] is not None else 0

        ss_min = int(mean_min_signal_strengh/len(item))
        ss_max = int(mean_max_signal_strengh/len(item))
        bat = int(mean_battery/len(item))
        m_a_l = list(mean_activity.values())[0]
        m_a = None if len(m_a_l) == 0 else sum(m_a_l)/len(m_a_l)
        mean_record = (epoch, epoch_f, mean_serial_number, ss_min if ss_min != 0 else None,
                       ss_max if ss_max != 0 else None, bat if bat != 0 else None, m_a)

        mean_data.append(mean_record)

        median_a = statistics.median(sorted(list(median_activity.values())))
        median_b = statistics.median(sorted(list(median_battery.values())))
        median_min_ss = statistics.median(sorted(list(median_min_signal_strengh.values())))
        median_max_ss = statistics.median(sorted(list(median_max_signal_strengh.values())))
        median_record = (epoch, epoch_f, median_serial_number, median_min_ss if median_min_ss != 0 else None,
                         median_max_ss if median_max_ss != 0 else None, median_b if median_b != 0 else None, median_a if median_b != 0 else None)
        median_data.append(median_record)

    result = mean_data + median_data

    return result


def process_raw_file(farm_id, data):
    start_time = time.time()
    farm_id = format_farm_id(farm_id)
    table_min, table_5min, table_10min, table_h, table_d, table_w, table_m = init_database(farm_id)
    print("process data for farm %s." % farm_id)

    # group records by animal id/serial number
    groups = defaultdict(list)
    for obj in data:
        groups[obj[2]].append(obj)
    animal_list_grouped_by_serialn = list(groups.values())

    data_resampled_min = []
    data_resampled_5min = []
    data_resampled_10min = []
    data_resampled_hour = []
    data_resampled_day = []
    data_resampled_week = []
    data_resampled_month = []

    MULTI_THREADING_ENABLED = True

    if MULTI_THREADING_ENABLED:
        pool = Pool(processes=5)

    animal_serial_number_to_ignore = [40101310107, 40101310146]
    for idx, animal_records in enumerate(animal_list_grouped_by_serialn):
        print("progress=%d/%d." % (idx, len(animal_list_grouped_by_serialn)))
        # find first and last record date.
        if len(animal_records) < 10000:
            continue
        serial_number = animal_records[0][2]
        if serial_number in animal_serial_number_to_ignore:
            continue
        first_timestamp, last_timestamp = get_first_last_timestamp(animal_records)
        if not MULTI_THREADING_ENABLED:
            data_resampled_min.extend(resample_to_min(first_timestamp, last_timestamp, animal_records))
            data_resampled_5min.extend(resample_to_5min(first_timestamp, last_timestamp, animal_records))
            data_resampled_10min.extend(resample_to_10min(first_timestamp, last_timestamp, animal_records))
            data_resampled_hour.extend(resample_to_hour(first_timestamp, last_timestamp, animal_records))
            data_resampled_day.extend(resample_to_day(first_timestamp, last_timestamp, animal_records))
            data_resampled_week.extend(resample_to_week(first_timestamp, last_timestamp, animal_records))
            data_resampled_month.extend(resample_to_month(first_timestamp, last_timestamp, animal_records))
        else:
            iterable = [animal_records]
            func1 = partial(resample_to_min, first_timestamp, last_timestamp)
            func2 = partial(resample_to_hour, first_timestamp, last_timestamp)
            func3 = partial(resample_to_day, first_timestamp, last_timestamp)
            func4 = partial(resample_to_week, first_timestamp, last_timestamp)
            func5 = partial(resample_to_month, first_timestamp, last_timestamp)
            func6 = partial(resample_to_10min, first_timestamp, last_timestamp)
            func7 = partial(resample_to_5min, first_timestamp, last_timestamp)
            result_func1 = pool.map_async(func1, iterable)
            result_func2 = pool.map_async(func2, iterable)
            result_func3 = pool.map_async(func3, iterable)
            result_func4 = pool.map_async(func4, iterable)
            result_func5 = pool.map_async(func5, iterable)
            result_func6 = pool.map_async(func6, iterable)
            result_func7 = pool.map_async(func7, iterable)
            r1 = result_func1.get()[0]
            r2 = result_func2.get()[0]
            r3 = result_func3.get()[0]
            r4 = result_func4.get()[0]
            r5 = result_func5.get()[0]
            r6 = result_func6.get()[0]
            r7 = result_func7.get()[0]
            data_resampled_min.extend(r1)
            data_resampled_hour.extend(r2)
            data_resampled_day.extend(r3)
            data_resampled_week.extend(r4)
            data_resampled_month.extend(r5)
            data_resampled_10min.extend(r6)
            data_resampled_5min.extend(r7)
    #save data in db
    print("saving data to db...")

    insert_m_record_to_sql_table_("%s_resolution_month" % farm_id, data_resampled_month)
    insert_m_record_to_sql_table_("%s_resolution_month" % farm_id, create_mean_median_animal(data_resampled_month))
    insert_m_record_to_sql_table_("%s_resolution_week" % farm_id, data_resampled_week)
    insert_m_record_to_sql_table_("%s_resolution_week" % farm_id, create_mean_median_animal(data_resampled_week))
    insert_m_record_to_sql_table_("%s_resolution_day" % farm_id, data_resampled_day)
    insert_m_record_to_sql_table_("%s_resolution_day" % farm_id, create_mean_median_animal(data_resampled_day))
    insert_m_record_to_sql_table_("%s_resolution_hour" % farm_id, data_resampled_hour)
    insert_m_record_to_sql_table_("%s_resolution_hour" % farm_id, create_mean_median_animal(data_resampled_hour))
    insert_m_record_to_sql_table_("%s_resolution_10min" % farm_id, data_resampled_10min)
    insert_m_record_to_sql_table_("%s_resolution_10min" % farm_id, create_mean_median_animal(data_resampled_10min))
    insert_m_record_to_sql_table_("%s_resolution_5min" % farm_id, data_resampled_5min)
    insert_m_record_to_sql_table_("%s_resolution_5min" % farm_id, create_mean_median_animal(data_resampled_5min))
    insert_m_record_to_sql_table("%s_resolution_min" % farm_id, data_resampled_min)
    insert_m_record_to_sql_table("%s_resolution_min" % farm_id, create_mean_median_animal_(data_resampled_min))
    create_indexes(farm_id)
    sql_db_flush()

    # if sys.argv[1] == 'h5':
    #     table_m.append(data_resampled_month)
    #     table_w.append(data_resampled_week)
    #     table_d.append(data_resampled_day)
    #     table_h.append(data_resampled_hour)
    #     table_min.append(data_resampled_min)
    #     table_10min.append(data_resampled_min)
    #     table_5min.append(data_resampled_min)
    #     table_m.flush()
    #     table_w.flush()
    #     table_d.flush()
    #     table_h.flush()
    #     table_min.flush()
    #     table_10min.flush()
    #     table_5min.flush()

    print(len(data_resampled_min), len(data_resampled_5min), len(data_resampled_10min), len(data_resampled_hour), len(data_resampled_day), len(data_resampled_week), len(data_resampled_month))
    print(get_elapsed_time_string(start_time, time.time()))
    print("finished processing raw file.")


def xl_date_to_date(xldate, wb):
    year, month, day, hour, minute, second = xlrd.xldate_as_tuple(xldate, wb.datemode)
    return "%02d/%02d/%d" % (day, month, year)


def convert_excel_time(cell_with_excel_time, wb):
    year, month, day, hour, minute, second = xlrd.xldate_as_tuple(cell_with_excel_time, wb.datemode)#year=month=day=0
    py_date = datetime(year=1989, month=4, day=12, hour=hour, minute=minute, second=second)#used dummy date we only need time
    time_value = py_date.strftime("%I:%M:%S %p")
    return time_value


def dump_cell(sheet, rowx, colx):
    c = sheet.cell(rowx, colx)
    xf = sheet.book.xf_list[c.xf_index]
    fmt_obj = sheet.book.format_map[xf.format_key]
    print(rowx, colx, repr(c.value), c.ctype, fmt_obj.type, fmt_obj.format_key, fmt_obj.format_str)
    return [rowx, colx, repr(c.value), c.ctype, fmt_obj.type, fmt_obj.format_key, fmt_obj.format_str]


def generate_raw_files_from_xlsx(directory_path, file_name):
    start_time = time.time()
    print("start readind xls files...")
    purge_file("log.txt")
    log_file = open("log.txt", "a")
    os.chdir(directory_path)
    file_paths = [val for sublist in
                  [[os.path.join(i[0], j) for j in i[2] if j.endswith('.xlsx')] for i in os.walk(directory_path)] for
                  val in sublist]
    print("founded %d files" % len(file_paths))
    print(file_paths)
    print("start generating raw file...")
    compression = False
    if compression:
        purge_file("raw_data_compressed_blosc_raw.h5")
        h5file = tables.open_file("raw_data_compressed_blosc_raw.h5", "w", driver="H5FD_CORE",
                                  filters=tables.Filters(complib='blosc', complevel=9))
    else:
        purge_file(file_name)
        # h5file = tables.open_file("raw_data.h5", mode="w", driver="H5FD_CORE")

        store = pandas.HDFStore(file_name)

    # table_f = h5file.create_table("/", "data", Animal, "Animal data in full resolution", expectedrows=33724492)
    # table_row = table_f.row
    valid_rows = 0

    for curr_file, path in enumerate(file_paths):
        # path = 'C:\\SouthAfrica\\Tracking Data\\Delmas\\June 2015\\70101200027_2015-5-31_06-00-00_to_2015-6-03_06-00-00.xlsx'
        # table_f = h5file.create_table("/", "data%d" % curr_file, Animal, path, expectedrows=33724492)
        # table_row = table_f.row
        df = []
        try:
            record_log = ""
            print("loading file in memory for reading...")
            print(path)
            book = xlrd.open_workbook(path)
            sheet = book.sheet_by_index(0)
            print("start reading...")
            found_col_index = False
            for row_index in xrange(0, sheet.nrows):
                try:
                    row_values = [sheet.cell(row_index, col_index).value for col_index in xrange(0, sheet.ncols)]
                    # print(path)
                    # print(row_values)
                    # find index of each column
                    if not found_col_index:
                        try:
                            date_col_index = row_values.index('Date')
                            time_col_index = row_values.index('Time')
                            control_station_col_index = row_values.index('Control station')
                            serial_number_col_index = row_values.index('Tag serial number')
                            signal_strength_col_index = row_values.index('Signal strength')
                            battery_voltage_col_index = row_values.index('Battery voltage')
                            first_sensor_value_col_index = row_values.index('First sensor value')
                            found_col_index = True
                        except ValueError:
                            date_col_index = 0
                            time_col_index = 1
                            control_station_col_index = 2
                            serial_number_col_index = 3
                            signal_strength_col_index = 4
                            battery_voltage_col_index = 5
                            first_sensor_value_col_index = 6

                    date_string = xl_date_to_date(row_values[date_col_index], book) + " " + convert_excel_time(
                        row_values[time_col_index], book)

                    epoch = int(datetime.strptime(date_string, '%d/%m/%Y %I:%M:%S %p').timestamp())
                    control_station = int(row_values[control_station_col_index])
                    serial_number = int(row_values[serial_number_col_index])
                    signal_strength = int(str(row_values[signal_strength_col_index]).replace("@", "").split('.')[0])
                    battery_voltage = int(str(row_values[battery_voltage_col_index]).split('.')[0], 16)
                    first_sensor_value = int(row_values[first_sensor_value_col_index])

                    record_log = "date_string=%s time=%s  row=%d epoch=%d control_station=%d serial_number=%d signal_strength=%d battery_voltage=%d first_sensor_value=%d" % (
                        date_string,
                        get_elapsed_time_string(start_time, time.time()), valid_rows, epoch, control_station,
                        serial_number,
                        signal_strength, battery_voltage, first_sensor_value)
                    # print(record_log)

                    df.append(
                        pandas.DataFrame({
                            'timestamp': epoch,
                            'control_station': control_station,
                            'serial_number': serial_number,
                            'signal_strength': signal_strength,
                            'battery_voltage': battery_voltage,
                            'first_sensor_value': first_sensor_value
                        }, index=[valid_rows]))

                    valid_rows += 1
                except Exception as exception:
                    print(exception)
                    print(path)
                    print(row_values)
                    log = "%d/%d--%s---%s---%s---%s" % (
                    curr_file, len(file_paths), get_elapsed_time_string(start_time, time.time()), str(exception), path,
                    record_log)
                    print(log)
                    log_file.write(log + "\n")
            del book
            del sheet
            store.append('/', value=pandas.concat(df), format='t', append=True,
                         data_columns=['timestamp', 'control_station', 'serial_number', 'signal_strength',
                                       'battery_voltage', 'first_sensor_value'])
            del df
            # del table_row
            # table_f.flush()
        except (Exception, FileNotFoundError, xlrd.biffh.XLRDError) as e:
            print(e)
            continue
    store.close()


if __name__ == '__main__':
    # generate_raw_files_from_xlsx("C:\SouthAfrica\Tracking Data\Delmas", "raw_data_delmas_debug.h5")
    print("start...")
    db_name = "south_africa"
    create_and_connect_to_sql_db(db_name)
    drop_all_tables(db_name)

    process_raw_h5files("E:\SouthAfrica\Tracking Data\\Delmas\\raw_data_delmas.h5")
